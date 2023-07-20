"""
This City Energy Analyst plugin is used to automate a number of parallelized simulations of the same scenario for a
single building, based on variable input stochastic distributions.
An output file is produced, which saves main inputs and outputs from each iteration.
"""

import numpy as np
import pandas as pd
import os
from openpyxl.reader.excel import load_workbook
from scipy.stats import beta
import multiprocessing
import csv

import cea.config
import cea.inputlocator
from cea.datamanagement import archetypes_mapper
import cea.resources.radiation.main as radiation
from cea.demand import demand_main
from cea.utilities.dbf import dbf_to_dataframe, dataframe_to_dbf
import cea.plugin

__author__ = "Moussa Geringswald"
__copyright__ = ""
__credits__ = [""]
__license__ = "MIT"
__version__ = "1.0"
__maintainer__ = ""
__email__ = "cea@arch.ethz.ch"
__status__ = "Production"

class ScenarioPlugin(cea.plugin.CeaPlugin):
    pass




def main(config):
    """
    This function contains the general inputs and parallelization.
    """
    assert os.path.exists(config.scenario), 'Scenario not found: %s' % config.scenario
    locator = cea.inputlocator.InputLocator(config.scenario, config.plugins)

    simulations = config.scenario_generator.iterations
   # def read_samples():
    samples_path = r"C:\Users\Moussa\Documents\BA E3D\Durchführung\plugin\samples\samples.csv"

    data = pd.read_csv(samples_path)


    """
    This function loads inputs from a dataframe into the CEA database, runs CEA scripts and stores inputs and outputs.
    """

    results = []
    # loop inside input dataframe to replace database for every iteration
    for i in range(simulations):
        print("Simulation number {}".format(i))
        n50 = data["n50"][i]
        U_wall = data["U_wall"][i]
        Cm_Af = data["Cm_Af"][i]



        types = load_workbook(filename=locator.get_database_construction_standards())
        envelope_types = types['ENVELOPE_ASSEMBLIES']
        envelope_types.cell(column=2, row=2).value = 'CONSTRUCTION_AS1'
        envelope_types.cell(column=3, row=2).value = 'TIGHTNESS_AS1'
        envelope_types.cell(column=6, row=2).value = 'WALL_AS1'
        envelope_types.cell(column=7, row=2).value = 'WALL_AS1'
        types.save(locator.get_database_construction_standards())

        # Changes and saves variables related to ENVELOPE
        assemblies_envelope = load_workbook(filename=locator.get_database_envelope_systems())
        construction = assemblies_envelope['CONSTRUCTION']
        construction.cell(column=3, row=2).value = Cm_Af
        tightness = assemblies_envelope['TIGHTNESS']
        tightness.cell(column=3, row=2).value = n50
        wall = assemblies_envelope['WALL']
        wall.cell(column=3, row=2).value = U_wall
        assemblies_envelope.save(locator.get_database_envelope_systems())


        ## Run CEA scripts: archetypes and energy demand

        config.multiprocessing = True  
        config.debug = False
        config.scenario = locator.scenario
        archetypes_mapper.main(config)  # loads database into the scenario
        radiation.main(config)         #runs radiation simulation
        demand_main.demand_calculation(locator, config)  # runs demand simulation


        # Extract relevant demand outputs
        Total_demand = pd.read_csv(locator.get_total_demand(), usecols=["Ea_kWh",
                                "El_kWh",
                                "Eve_kWh",
                                "Ev_kWh",
                                "Edata_kWh",
                                "Epro_kWh",
                                "Eaux_kWh",
                                "Qhs_sys_kWh",
                                "Qww_sys_kWh",
                                "Qcs_sys_kWh",
                                'Qcdata_sys_kWh',
                                'Qcre_sys_kWh'
                                ])
        Heat_Load = Total_demand[][]+

        # Storage of outputs in a dict
        dict_outputs = {
            "Annual_energy_demand_MWhyr": round(Annual_energy_demand_MWhyr, 2)
        }

        # Convert outputs into a csv file
        results.append(dict_outputs)
        df = pd.DataFrame(results)
        df.to_csv(r"C:\Users\Moussa\Documents\BA E3D\Durchführung\plugin\results\results.csv", index=False)


    print("Demand samples have been simulated")




if __name__ == '__main__':
    main(cea.config.Configuration())
